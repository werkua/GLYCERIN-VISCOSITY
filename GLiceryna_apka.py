import serial
import openpyxl
from datetime import datetime, timedelta

def read_data_from_serial(ser):
    data = ser.readline().decode('utf-8').rstrip()
    return data

def get_ball_info(ball_number):
    print(f"Ball {ball_number} information:")
    mass = input("Enter mass (in grams): ")
    diameter = input("Enter diameter (in millimeters): ")
    return mass, diameter

def main():
    # Open the connection to the Arduino's serial port
    ser = serial.Serial('COM3', 9600)  # Replace 'COM3' with the appropriate port for your Arduino
    print("Connected to Arduino")

    # Create a new Excel workbook and sheet
    wb = openpyxl.Workbook()
    ws = wb.active

    # Add column headers
    ws.append(["Ball Number", "Mass (g)", "Diameter (mm)", "Time (s)", "Measurement Time", "Additional Comments"])

    balls_data = {}
    last_ball_number = 0

    try:
        filename = "Glycerin_" + datetime.now().strftime("%Y%m%d_%H%M%S") + ".xlsx"

        while True:
            # Read data from Arduino (time in milliseconds separated by a comma)
            data = read_data_from_serial(ser)
            if not data:
                continue

            # Convert the time to seconds as a floating-point number
            try:
                time_in_seconds = float(data)/1000
            except ValueError:
                print(f"Invalid time data received: {data}")
                continue

            # Calculate the measurement time based on the time in seconds
            base_time = datetime.now() - timedelta(seconds=time_in_seconds)
            formatted_time = base_time.strftime("%Y-%m-%d %H:%M:%S")  # Time format as "YYYY-MM-DD HH:MM:SS"

            # Ask the user if they want to add a new ball or choose an existing one
            user_choice = input("Add new ball (n) or choose from existing (c)? ").lower()

            if user_choice == "n":
                ball_number = last_ball_number + 1
                mass, diameter = get_ball_info(ball_number)
                balls_data[ball_number] = (mass, diameter)
                ws.append([ball_number, mass, diameter, time_in_seconds, formatted_time, ""])
                last_ball_number += 1
            elif user_choice == "c":
                ball_number = int(input("Enter ball number: "))
                if ball_number not in balls_data:
                    print(f"Ball {ball_number} not found in the data. Please add the ball first.")
                    continue
                mass, diameter = balls_data[ball_number]
                ws.append([ball_number, mass, diameter, time_in_seconds, formatted_time, ""])
            else:
                print("Invalid choice. Please enter 'n' or 'c'.")
                continue

            # Ask user for additional comments
            additional_comments = input("Enter additional comments: ")
            ws.cell(row=len(ws["A"]) - 1, column=6, value=additional_comments)

            # Save data to the Excel file
            wb.save(filename)
            print(f"Ball Number: {ball_number}, Mass (g): {mass}, Diameter (mm): {diameter}, Time (s): {time_in_seconds}, Measurement Time: {formatted_time}, Additional Comments: {additional_comments}")

    except KeyboardInterrupt:
        # Handle user interruption (e.g., CTRL+C)
        print("Interrupted by the user")

    except Exception as e:
        # Handle other errors
        print("An error occurred:", e)

    finally:
        # Close the connection to Arduino
        ser.close()
        print("Closed the connection to Arduino")

if __name__ == "__main__":
    main()
